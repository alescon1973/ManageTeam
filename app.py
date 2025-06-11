import streamlit as st
import pandas as pd
import os
import calendar
import json
from datetime import datetime, time
from streamlit_extras.add_vertical_space import add_vertical_space
import locale
from openpyxl import load_workbook

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from io import BytesIO

import plotly.express as px
import plotly.graph_objects as go


import logging
import glob



# Imposta la localizzazione italiana per i nomi dei mesi

# Percorsi base
dir_squadre = "squadre"
dir_presenze = "presenze"
nomi_squadre = ["PP", "U19", "U18", "U17R", "U17P", "U16R", "U16P", "U15R", "U15P", "U14R", "U14P"]

# Crea directory se non esistono
os.makedirs(dir_squadre, exist_ok=True)
os.makedirs(dir_presenze, exist_ok=True)

# Funzione per caricare la squadra
def load_squad(squadra):
    path = os.path.join(dir_squadre, f"{squadra}.csv")
    if os.path.exists(path):
        return pd.read_csv(path, sep=';')
    else:
        return pd.DataFrame(columns=["NOME", "COGNOME", "ANNO", "RUOLO"])

# Funzione per salvare la squadra
def save_squad(squadra, df):
    path = os.path.join(dir_squadre, f"{squadra}.csv")
    df.to_csv(path, sep=';', index=False)

# Funzione per ottenere descrizione squadra
def get_squadra_descrizione(codice):
    if codice == "PP":
        return "Prima Squadra"
    elif codice.startswith("U"):
        livello = ""
        if codice.endswith("R"):
            livello = "Regionale"
        elif codice.endswith("P"):
            livello = "Provinciale"
        categoria = codice[1:3]  # U19 -> 19
        return f"Under {categoria} {livello}".strip()
    return codice

# Funzioni per gestire presenze JSON
def get_presenze_path(squadra):
    return os.path.join(dir_presenze, f"{squadra}.json")

def load_presenze(squadra):
    path = get_presenze_path(squadra)
    if os.path.exists(path):
        with open(path, "r") as f:
            return json.load(f)
    return {}

def save_presenze(squadra, data):
    path = get_presenze_path(squadra)
    with open(path, "w") as f:
        json.dump(data, f, indent=2)

# Funzione per esportare in Excel
def salva_excel_convocazione(dir_path, squadra_sel, squadra_avversaria, data_incontro, ora_incontro, campo, ora_raduno, convocati, non_convocati, mister, dirigente):
    modello_path = "Convocazione.xlsx"
    if not os.path.exists(modello_path):
        st.error("File modello Convocazione.xlsx non trovato nella root del progetto.")
        return

    wb = load_workbook(modello_path)
    ws = wb.active

    ws["C10"] = squadra_avversaria
    ws["C14"] = f"{data_incontro.strftime('%d/%m/%Y')} - {ora_incontro.strftime('%H:%M')}"

    if "," in campo:
        parte1, parte2 = campo.split(",", 1)
        ws["C16"] = parte1.strip()
        ws["C17"] = parte2.strip()
    else:
        ws["C16"] = campo.strip()
        ws["C17"] = ""

    ws["C19"] = ora_raduno

    for idx, giocatore in enumerate(convocati[:22]):
        ws[f"C{22 + idx}"] = giocatore

    ws["C45"] = non_convocati
    ws["C50"] = mister
    ws["C52"] = dirigente

    dir_squadra = os.path.join(dir_path, squadra_sel)
    os.makedirs(dir_squadra, exist_ok=True)
    nome_file = f"Convocazione_{squadra_sel}_{squadra_avversaria.replace(' ', '_')}.xlsx"
    percorso_file = os.path.join(dir_squadra, nome_file)
    wb.save(percorso_file)

    with open(percorso_file, "rb") as f:
        st.download_button("Scarica convocazione in Excel", data=f, file_name=nome_file, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def genera_pdf_partita(dati, logo_path="logo.png"):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 50

    # Inserisci logo in alto a sinistra se esiste
    if os.path.exists(logo_path):
        try:
            c.drawImage(ImageReader(logo_path), 40, height - 100, width=80, preserveAspectRatio=True, mask='auto')
        except Exception as e:
            print("Errore caricamento logo:", e)

    y -= 100  # Spazio sotto al logo

    def scrivi_riga(testo, gap=16, bold=False):
        nonlocal y
        if y < 100:
            c.showPage()
            y = height - 50
        if bold:
            c.setFont("Helvetica-Bold", 11)
        else:
            c.setFont("Helvetica", 10)
        c.drawString(50, y, testo)
        y -= gap

    scrivi_riga(f"📄 REPORT PARTITA", bold=True)
    scrivi_riga(f"Giornata: {dati['giornata']} - Squadra: {dati['squadra']}")
    scrivi_riga(f"In casa: {dati['home_away']} - Risultato: {dati['risultato']} - Recupero: {dati['recupero']} min")
    scrivi_riga("-" * 90)

    scrivi_riga("🧤 Formazione:", bold=True)
    for i, giocatore in enumerate(dati["formazione"], 1):
        ruolo = "Titolare" if i <= 11 else "Panchina"
        scrivi_riga(f"{i}. {giocatore} ({ruolo})")
    scrivi_riga("-" * 90)

    if dati["substitutions"]:
        scrivi_riga("🔁 Sostituzioni:", bold=True)
        for s in dati["substitutions"]:
            scrivi_riga(f"Min {s['time_sub']}: {s['sub_out']} → {s['sub_in']}")
        scrivi_riga("-" * 90)

    if dati["goal"]:
        scrivi_riga("⚽ Gol:", bold=True)
        for i, g in enumerate(dati["goal"], 1):
            scrivi_riga(f"{i}. {g}")
        scrivi_riga("-" * 90)

    if dati["ammonizioni"]:
        scrivi_riga("🟨 Ammoniti:", bold=True)
        for a in dati["ammonizioni"]:
            scrivi_riga(f"- {a}")
        scrivi_riga("-" * 90)

    if dati["espulsioni"]:
        scrivi_riga("🟥 Espulsioni:", bold=True)
        for e in dati["espulsioni"]:
            scrivi_riga(f"- {e['esp_player']} al minuto {e['time_esp']}")
        scrivi_riga("-" * 90)

    if dati["non_convocati"]:
        scrivi_riga("🚫 Non convocati:", bold=True)
        for nc in dati["non_convocati"]:
            scrivi_riga(f"- {nc['giocatore']} ({nc['motivo']})")
        scrivi_riga("-" * 90)

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer



# Pagina iniziale
st.set_page_config(page_title="Gestione Squadre Giovanili", layout="wide")

if "pagina" not in st.session_state:
    st.session_state.pagina = "home"

if "sezione" not in st.session_state:
    st.session_state.sezione = "Squadra"

if st.session_state.pagina == "home":
    col_logo, col_titolo = st.columns([1, 5])
    with col_logo:
        st.image("static/logo.png", width=120)
    with col_titolo:
        st.title("Athletic Soccer Academy")
    st.header("Seleziona una squadra per iniziare")
    squadra_sel = st.selectbox("Scegli la squadra", nomi_squadre)
    if st.button("Vai alla dashboard"):
        st.session_state.squadra_sel = squadra_sel
        st.session_state.pagina = "dashboard"
        st.rerun()

elif st.session_state.pagina == "dashboard":
    squadra_sel = st.session_state.squadra_sel
    squadra_descrizione = get_squadra_descrizione(squadra_sel)
    df = load_squad(squadra_sel)

    with st.sidebar:
        st.title("Menu")
        st.markdown(f"**Squadra selezionata:** {squadra_sel}")

        if st.button("👥 Squadra"):  # gestione rosa
            st.session_state.sezione = "Squadra"
            st.rerun()

        if st.button("🗓️ Presenze"):  # registro presenze
            st.session_state.sezione = "Presenze"
            st.rerun()

        if st.button("📣 Convocazioni"):  # nuova convocazione
            st.session_state.sezione = "Convocazioni"
            st.rerun()

        if st.button("✏️ Modifica Convocazione"):  # modifica convocazione
            st.session_state.sezione = "Modifica Convocazione"
            st.rerun()

        if st.button("🎮 Partita"):  # gestione match
            st.session_state.sezione = "Partita"
            st.rerun()

        if st.button("📊 Reportistica"):
            st.session_state.sezione = "Reportistica"
            st.rerun()


        add_vertical_space(1)
        if st.button("🔙 Torna alla selezione squadra"):
            st.session_state.pagina = "home"
            st.rerun()


    # Layout logo + titolo
    col1, col2 = st.columns([1, 4])
    with col1:
        st.image("static/logo.png", width=100)
    with col2:
        st.title("Athletic Soccer Academy")
        st.subheader(f"{squadra_sel} - {squadra_descrizione}")
    
    st.header(st.session_state.sezione)


    if st.session_state.sezione == "Squadra":
        edited_df = st.data_editor(
            df.copy(),
            num_rows="dynamic",
            use_container_width=True,
            key=f"data_editor_modifica_{squadra_sel}",
            disabled=False
        )

        if st.button("Salva lista giocatori"):
            save_squad(squadra_sel, edited_df)
            st.success("Lista salvata con successo!")
            st.rerun()

    elif st.session_state.sezione == "Presenze":
        col1, col2 = st.columns(2)

        # Nomi mesi italiani hardcoded per evitare locale
        mesi_italiani = [
            "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
            "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"
        ]

        with col1:
            mese = st.selectbox("Mese", mesi_italiani)
        with col2:
            anno = st.number_input("Anno", min_value=2020, max_value=2100, value=datetime.now().year)

        mese_numero = mesi_italiani.index(mese) + 1
        giorni_mese = calendar.monthrange(anno, mese_numero)[1]
        date_colonne = [f"{giorno:02d}/{mese_numero:02d}" for giorno in range(1, giorni_mese + 1)]

        nomi_giocatori = df[["NOME", "COGNOME"]].apply(lambda x: f"{x['NOME']} {x['COGNOME']}", axis=1).tolist()
        presenze_data = load_presenze(squadra_sel)
        chiave_mese = f"{anno}-{mese_numero:02d}"

        if chiave_mese in presenze_data:
            data_presenze = pd.DataFrame(presenze_data[chiave_mese])
        else:
            data_presenze = pd.DataFrame(index=nomi_giocatori, columns=date_colonne)

        data_presenze = data_presenze.reindex(index=nomi_giocatori, columns=date_colonne)
        data_presenze = data_presenze.where(pd.notnull(data_presenze), "")

        edited_df = st.data_editor(
            data_presenze,
            num_rows="dynamic",
            use_container_width=True,
            key=f"presenze_{chiave_mese}",
            column_config={
                col: st.column_config.SelectboxColumn(
                    label=col,
                    options=["", "P", "AI", "MS", "ML", "I", "MP"]
                ) for col in data_presenze.columns
            },
        )

        if st.button("Salva presenze"):
            presenze_data[chiave_mese] = edited_df.where(pd.notnull(edited_df), "").to_dict()
            save_presenze(squadra_sel, presenze_data)
            st.success("Presenze salvate correttamente.")


    elif st.session_state.sezione == "Convocazioni":
        # Crea directory convocazioni per la squadra se non esiste
        dir_convocazioni_squadra = os.path.join("convocazioni", squadra_sel)
        os.makedirs(dir_convocazioni_squadra, exist_ok=True)
        
        # st.subheader("Nuova Convocazione")
        
        # Form per i dettagli della partita
        col1, col2 = st.columns(2)
        with col1:
            giornata = st.number_input("Giornata", min_value=1, max_value=50, value=1)
            squadra_avversaria = st.text_input("Squadra avversaria")
            data_incontro = st.date_input("Data incontro", value=datetime.now().date())
            ora_incontro = st.time_input("Ora incontro", value=datetime.now().time().replace(second=0, microsecond=0))
            data_ora_incontro = datetime.combine(data_incontro, ora_incontro).strftime("%Y-%m-%dT%H:%M")


        with col2:
            denominazione_campo = st.text_input("Denominazione campo")
            ora_raduno = st.text_input("Ora raduno (HH:MM)", value="13:15")
            nome_mister = st.text_input("Mister")
            nome_dirigente = st.text_input("Dirigente")
        
        st.markdown("---")
        st.subheader("Selezione Giocatori")
        
        # Carica la lista dei giocatori
        df_squadra = load_squad(squadra_sel)
        nomi_giocatori = df_squadra[["NOME", "COGNOME"]].apply(lambda x: f"{x['NOME']} {x['COGNOME']}", axis=1).tolist()
        
        # Layout a due colonne
        col_disponibili, col_convocati = st.columns([1, 1])
        
        # Inizializza la lista dei convocati nella session state
        if 'convocati' not in st.session_state:
            st.session_state.convocati = [""] * 20
        
        # Lista dei giocatori disponibili
        with col_disponibili:
            st.markdown("**Giocatori disponibili**")
            convocati_attuali = [p for p in st.session_state.convocati if p]
            disponibili = [g for g in nomi_giocatori if g not in convocati_attuali]

            for giocatore in disponibili:
                if st.button(giocatore, key=f"disp_{giocatore}"):
                    for i in range(len(st.session_state.convocati)):
                        if st.session_state.convocati[i] == "":
                            st.session_state.convocati[i] = giocatore
                            st.rerun()


        
        with col_convocati:
            st.markdown("**Giocatori convocati (max 20)**")

            if "convocati" not in st.session_state:
                st.session_state.convocati = [""] * 20

            # Salva una copia dello stato iniziale per confronto
            vecchi_convocati = st.session_state.convocati.copy()

            for i in range(20):
                current_player = st.session_state.convocati[i]

                # Calcola occupati: tutti tranne quello corrente
                occupati = [st.session_state.convocati[j] for j in range(20) if j != i and st.session_state.convocati[j]]

                # Opzioni disponibili: tutti i giocatori meno quelli occupati
                opzioni = [""] + [p for p in nomi_giocatori if p not in occupati or p == current_player]

                cols = st.columns([0.1, 0.9])
                with cols[0]:
                    st.markdown(f"**{i+1}.**")
                with cols[1]:
                    selezionato = st.selectbox(
                        "", options=opzioni,
                        index=opzioni.index(current_player) if current_player in opzioni else 0,
                        key=f"conv_select_{i}",
                        label_visibility="collapsed"
                    )

                    # Aggiorna lo stato solo se c'è stato un cambiamento
                    if selezionato != current_player:
                        st.session_state.convocati[i] = selezionato
                        st.rerun()  # Rerun immediato per ricalcolare tutti gli slot
                
        # Calcola i non convocati
        non_convocati = [g for g in nomi_giocatori if g not in st.session_state.convocati]
        
        # Textarea per i non convocati
        st.subheader("Giocatori non convocati")
        non_convocati_text = st.text_area(
            "Non convocati (separati da virgola)", 
            value=", ".join(non_convocati),
            height=100
        )
        
        # Pulsante salva
        if st.button("Salva Convocazione"):
            if not squadra_avversaria:
                st.error("Inserire il nome della squadra avversaria")
            else:
                # Prepara i dati da salvare
                convocazione_data = {
                    "giornata": giornata,
                    "squadra": squadra_sel,
                    "squadra_avversaria": squadra_avversaria,
                    "data_ora_incontro": data_ora_incontro,
                    "denominazione_campo": denominazione_campo,
                    "ora_raduno": ora_raduno,
                    "componenti_squadra": [p for p in st.session_state.convocati if p],
                    "non_convocati": non_convocati_text,
                    "nome_mister": nome_mister,
                    "nome_dirigente": nome_dirigente
                }
                
                # Salva il file
                filename = f"{giornata}_{squadra_avversaria.replace(' ', '_')}.json"
                filepath = os.path.join(dir_convocazioni_squadra, filename)
                
                with open(filepath, "w") as f:
                    json.dump(convocazione_data, f, indent=2, ensure_ascii=False)
                st.success(f"Convocazione salvata correttamente in {filename}")

                # Salvataggio Excel + download
                salva_excel_convocazione(
                    dir_path = dir_convocazioni_squadra,
                    squadra_sel=squadra_sel,
                    squadra_avversaria=squadra_avversaria,
                    data_incontro=data_incontro,
                    ora_incontro=ora_incontro,
                    campo=denominazione_campo,
                    ora_raduno=ora_raduno,
                    convocati=[p for p in st.session_state.convocati if p],
                    non_convocati=non_convocati_text,
                    mister=nome_mister,
                    dirigente=nome_dirigente
                )

    # Sezione per modificare una convocazione esistente
    elif st.session_state.sezione == "Modifica Convocazione":
        #st.subheader("Modifica Convocazione Esistente")

        dir_convocazioni_squadra = os.path.join("convocazioni", squadra_sel)

        convocazioni_esistenti = [
            f for f in os.listdir(dir_convocazioni_squadra)
            if f.endswith(".json")
        ]

        # Ordina i file per numero di giornata all'inizio del nome
        def estrai_numero(file):
            try:
                return int(file.split("_")[0])
            except ValueError:
                return 0

        convocazioni_esistenti.sort(key=estrai_numero)

        if not convocazioni_esistenti:
            st.info("Nessuna convocazione salvata per questa squadra.")
        else:
            file_scelto = st.selectbox("Seleziona convocazione da modificare", [""] + convocazioni_esistenti)

            if file_scelto:  # Mostra il resto solo se un file è selezionato
                percorso = os.path.join(dir_convocazioni_squadra, file_scelto)

                with open(percorso, "r") as f:
                    dati = json.load(f)

                data_str, ora_str = dati["data_ora_incontro"].split("T")
                data_incontro = datetime.strptime(data_str, "%Y-%m-%d").date()
                ora_incontro = datetime.strptime(ora_str, "%H:%M").time()

                st.markdown("---")
                st.subheader("Dettagli Incontro")

                col1, col2 = st.columns(2)

                with col1:
                    giornata = st.number_input("Giornata", min_value=1, max_value=50, value=dati["giornata"])
                    squadra_avversaria = st.text_input("Squadra avversaria", value=dati["squadra_avversaria"])
                    data_incontro = st.date_input("Data incontro", value=data_incontro)
                    ora_incontro = st.time_input("Ora incontro", value=ora_incontro)

                with col2:
                    denominazione_campo = st.text_input("Denominazione campo", value=dati["denominazione_campo"])
                    ora_raduno = st.text_input("Ora raduno (HH:MM)", value=dati["ora_raduno"])
                    nome_mister = st.text_input("Mister", value=dati["nome_mister"])
                    nome_dirigente = st.text_input("Dirigente", value=dati["nome_dirigente"])



                st.markdown("---")
                st.subheader("Selezione Giocatori")

                df_squadra = load_squad(squadra_sel)
                nomi_giocatori = df_squadra[["NOME", "COGNOME"]].apply(lambda x: f"{x['NOME']} {x['COGNOME']}", axis=1).tolist()

                col_disponibili, col_convocati = st.columns([1, 1])

                convocati_caricati = dati.get("componenti_squadra", [])
                if 'convocati' not in st.session_state or file_scelto != st.session_state.get("convocazione_corrente"):
                    st.session_state.convocati = convocati_caricati + [""] * (20 - len(convocati_caricati))
                    st.session_state.convocazione_corrente = file_scelto

                with col_disponibili:
                    st.markdown("**Giocatori disponibili**")
                    convocati_attuali = [p for p in st.session_state.convocati if p]
                    disponibili = [g for g in nomi_giocatori if g not in convocati_attuali]

                    for giocatore in disponibili:
                        if st.button(giocatore, key=f"disp_mod_{giocatore}"):
                            for i in range(len(st.session_state.convocati)):
                                if st.session_state.convocati[i] == "":
                                    st.session_state.convocati[i] = giocatore
                                    st.rerun()

                with col_convocati:
                    st.markdown("**Giocatori convocati (max 20)**")
                    for i in range(20):
                        current_player = st.session_state.convocati[i]
                        occupati = [st.session_state.convocati[j] for j in range(20) if j != i and st.session_state.convocati[j]]
                        opzioni = [""] + [p for p in nomi_giocatori if p not in occupati or p == current_player]

                        cols = st.columns([0.1, 0.9])
                        with cols[0]:
                            st.markdown(f"**{i+1}.**")
                        with cols[1]:
                            selezionato = st.selectbox(
                                "", options=opzioni,
                                index=opzioni.index(current_player) if current_player in opzioni else 0,
                                key=f"conv_select_mod_{i}",
                                label_visibility="collapsed"
                            )

                            if selezionato != current_player:
                                st.session_state.convocati[i] = selezionato
                                st.rerun()

                st.subheader("Giocatori non convocati")

                convocati_attuali = [p for p in st.session_state.convocati if p]
                non_convocati = [g for g in nomi_giocatori if g not in convocati_attuali]
                non_convocati_text = st.text_area(
                    "Non convocati (separati da virgola)",
                    value=", ".join(non_convocati),
                    height=100
                )

                if st.button("Salva modifiche convocazione"):
                    nuovi_dati = {
                        "giornata": giornata,
                        "squadra": squadra_sel,
                        "squadra_avversaria": squadra_avversaria,
                        "data_ora_incontro": data_ora_incontro,
                        "denominazione_campo": denominazione_campo,
                        "ora_raduno": ora_raduno,
                        "componenti_squadra": [p for p in st.session_state.convocati if p],
                        "non_convocati": non_convocati_text,
                        "nome_mister": nome_mister,
                        "nome_dirigente": nome_dirigente
                    }

                    with open(percorso, "w") as f:
                        json.dump(nuovi_dati, f, indent=2, ensure_ascii=False)

                    st.success("Convocazione modificata con successo!")

                    salva_excel_convocazione(
                        squadra_sel=squadra_sel,
                        squadra_avversaria=squadra_avversaria,
                        data_incontro=data_incontro,
                        ora_incontro=ora_incontro,
                        campo=denominazione_campo,
                        ora_raduno=ora_raduno,
                        convocati=[p for p in st.session_state.convocati if p],
                        non_convocati=non_convocati_text,
                        mister=nome_mister,
                        dirigente=nome_dirigente
                    )



    elif st.session_state.sezione == "Partita":
        st.subheader("Gestione Partita")

        dir_conv_squadra = os.path.join("convocazioni", squadra_sel)
        convocazioni = sorted([f for f in os.listdir(dir_conv_squadra) if f.endswith(".json")])

        file_conv = st.selectbox("Seleziona convocazione", [""] + convocazioni, index=0)

        if file_conv:
            with open(os.path.join(dir_conv_squadra, file_conv), "r") as f:
                dati_conv = json.load(f)

            # Tutto il contenuto della scheda parte da qui
            st.markdown("---")
            st.subheader("Dati partita")

            giornata = dati_conv["giornata"]
            squadra_avversaria = dati_conv["squadra_avversaria"]
            giocatori_convocati = dati_conv["componenti_squadra"]

            col1, col2 = st.columns(2)
            with col1:
                st.text_input("Giornata", value=str(giornata), disabled=True)
                st.text_input("Squadra avversaria", value=squadra_avversaria, disabled=True)
                home_away = st.selectbox("Casa/Fuori", ["Casa", "Fuori casa"])
                risultato = st.text_input("Risultato (es. 2-1, lo scrivi come se giocassi sempre in casa)")
            with col2:
                recupero = st.number_input("Minuti di recupero", min_value=0, max_value=20, value=5)

            st.markdown("### Formazione")

            if "formazione" not in st.session_state:
                st.session_state.formazione = [""] * 20

            col_disp, col_form = st.columns([1, 1])
            with col_disp:
                st.markdown("**Convocati disponibili**")
                occupati = [g for g in st.session_state.formazione if g]
                disponibili = [g for g in giocatori_convocati if g not in occupati]
                for g in disponibili:
                    if st.button(g, key=f"disp_partita_{g}"):
                        for i in range(20):
                            if st.session_state.formazione[i] == "":
                                st.session_state.formazione[i] = g
                                st.rerun()

            with col_form:
                st.markdown("**Formazione**")
                for i in range(20):
                    current = st.session_state.formazione[i]
                    opzioni = [""] + [g for g in giocatori_convocati if g not in st.session_state.formazione or g == current]
                    cols = st.columns([0.1, 0.9])
                    with cols[0]:
                        st.markdown(f"{i+1}.")
                    with cols[1]:
                        selezionato = st.selectbox("", opzioni, index=opzioni.index(current) if current in opzioni else 0,
                                                key=f"form_slot_{i}", label_visibility="collapsed")
                        if selezionato != current:
                            st.session_state.formazione[i] = selezionato
                            st.rerun()

            st.markdown("---")

            st.markdown("### Sostituzioni")

            # Reset sostituzioni
            if st.button("🔄 Reset sostituzioni"):
                st.session_state.sostituzioni = []
                st.rerun()

            col_sost, _ = st.columns([1, 5])
            with col_sost:
                num_sost = st.number_input(
                    "Numero di sostituzioni",
                    min_value=0, max_value=9999, step=1,
                    label_visibility="visible"
    )
            if "sostituzioni" not in st.session_state or len(st.session_state.sostituzioni) != num_sost:
                st.session_state.sostituzioni = [{"in": "", "out": "", "minuto": ""} for _ in range(num_sost)]

            # Formazione dinamica aggiornata dopo ogni sostituzione
            formazione_dinamica = st.session_state.formazione[:11]
            panchina_dinamica = st.session_state.formazione[11:]

            for i in range(num_sost):
                st.markdown(f"**Sostituzione {i+1}**")

                # Disponibili aggiornati: nessun doppione
                disponibili_in = [p for p in panchina_dinamica if p and p not in [s["in"] for s in st.session_state.sostituzioni[:i]]]
                disponibili_out = [p for p in formazione_dinamica if p and p not in [s["out"] for s in st.session_state.sostituzioni[:i]]]

                col_in, col_out, col_min = st.columns(3)

                with col_in:
                    st.session_state.sostituzioni[i]["in"] = st.selectbox(
                        "Entra", [""] + disponibili_in,
                        index=([""] + disponibili_in).index(st.session_state.sostituzioni[i]["in"]) if st.session_state.sostituzioni[i]["in"] in disponibili_in else 0,
                        key=f"sost_in_{i}"
                    )

                with col_out:
                    st.session_state.sostituzioni[i]["out"] = st.selectbox(
                        "Esce", [""] + disponibili_out,
                        index=([""] + disponibili_out).index(st.session_state.sostituzioni[i]["out"]) if st.session_state.sostituzioni[i]["out"] in disponibili_out else 0,
                        key=f"sost_out_{i}"
                    )

                with col_min:
                    min_prev = int(st.session_state.sostituzioni[i-1]["minuto"]) if i > 0 and st.session_state.sostituzioni[i-1]["minuto"].isdigit() else 0
                    current_val = st.session_state.sostituzioni[i]["minuto"]
                    nuovo_minuto = st.text_input("Minuto", value=current_val, key=f"sost_min_{i}")
                    if nuovo_minuto.isdigit() and int(nuovo_minuto) < min_prev:
                        st.warning(f"Il minuto della sostituzione {i+1} deve essere ≥ {min_prev}")
                    st.session_state.sostituzioni[i]["minuto"] = nuovo_minuto

                # Applica dinamicamente la sostituzione alla formazione
                in_player = st.session_state.sostituzioni[i]["in"]
                out_player = st.session_state.sostituzioni[i]["out"]
                if in_player and out_player and out_player in formazione_dinamica:
                    formazione_dinamica = [p if p != out_player else in_player for p in formazione_dinamica]
                    panchina_dinamica = [p for p in panchina_dinamica if p != in_player]

            st.markdown("---")

            st.markdown("### Ammonizioni")

            if "ammoniti" not in st.session_state:
                st.session_state.ammoniti = []

            col_amm, _ = st.columns([1, 5])
            with col_amm:
                num_ammoniti = st.number_input(
                    "Numero di giocatori ammoniti",
                    min_value=0, max_value=11, step=1,
                    label_visibility="visible")


            # Reset automatico se cambia il numero
            if len(st.session_state.ammoniti) != num_ammoniti:
                st.session_state.ammoniti = [""] * num_ammoniti

            disponibili_ammoniti = [g for g in giocatori_convocati if g not in st.session_state.ammoniti]

            for i in range(num_ammoniti):
                opzioni = [""] + [g for g in giocatori_convocati if g not in st.session_state.ammoniti or g == st.session_state.ammoniti[i]]
                st.session_state.ammoniti[i] = st.selectbox(
                    f"Ammonito {i+1}",
                    options=opzioni,
                    index=opzioni.index(st.session_state.ammoniti[i]) if st.session_state.ammoniti[i] in opzioni else 0,
                    key=f"ammonito_{i}"
                )

            st.markdown("---")

            st.markdown("### Espulsioni")

            if "espulsioni" not in st.session_state:
                st.session_state.espulsioni = []

            col_esp, _ = st.columns([1, 5])
            with col_esp:
                num_espulsioni = st.number_input(
                    "Numero di espulsioni",
                    min_value=0, max_value=11, step=1,
                    label_visibility="visible")

            # Inizializza/resetta se cambia il numero
            if len(st.session_state.espulsioni) != num_espulsioni:
                st.session_state.espulsioni = [{"giocatore": "", "minuto": ""} for _ in range(num_espulsioni)]

            for i in range(num_espulsioni):
                st.markdown(f"**Espulsione {i+1}**")
                espulsi_precedenti = [e["giocatore"] for e in st.session_state.espulsioni[:i]]
                opzioni = [""] + [g for g in giocatori_convocati if g not in espulsi_precedenti or g == st.session_state.espulsioni[i]["giocatore"]]

                col_gioc, col_min = st.columns(2)

                with col_gioc:
                    st.session_state.espulsioni[i]["giocatore"] = st.selectbox(
                        "Espulso",
                        options=opzioni,
                        index=opzioni.index(st.session_state.espulsioni[i]["giocatore"]) if st.session_state.espulsioni[i]["giocatore"] in opzioni else 0,
                        key=f"espulsione_gioc_{i}"
                    )

                with col_min:
                    st.session_state.espulsioni[i]["minuto"] = st.text_input(
                        "Minuto",
                        value=st.session_state.espulsioni[i]["minuto"],
                        key=f"espulsione_min_{i}"
                    )
            
            st.markdown("---")

            st.markdown("### Gol")

            # Estrai numero gol fatti (prima del "-")
            try:
                gol_fatti = int(risultato.split("-")[0].strip())
            except (IndexError, ValueError):
                gol_fatti = 0

            if "gol" not in st.session_state or len(st.session_state.gol) != gol_fatti:
                st.session_state.gol = [""] * gol_fatti

            opzioni_gol = ["autogol"] + giocatori_convocati

            for i in range(gol_fatti):
                st.session_state.gol[i] = st.selectbox(
                    f"Gol {i+1} - Marcatore",
                    options=opzioni_gol,
                    index=opzioni_gol.index(st.session_state.gol[i]) if st.session_state.gol[i] in opzioni_gol else 0,
                    key=f"gol_{i}"
                )

            st.markdown("---")
            st.markdown("### Giocatori non convocati e motivazioni")

            motivi_disponibili = [
                "Scelta tecnica",
                "Infortunato",
                "Squalificato",
                "Altra categoria",
                "Indisponibile",
                "Non allenato",
                "Allenamenti insufficienti",
                "Malattia"
            ]

            # Estrai lista dei non convocati dalla stringa
            non_convocati_lista = [
                nome.strip() for nome in dati_conv.get("non_convocati", "").split(",") if nome.strip()
            ]

            if "motivi_non_convocati" not in st.session_state or len(st.session_state.motivi_non_convocati) != len(non_convocati_lista):
                st.session_state.motivi_non_convocati = {nome: "" for nome in non_convocati_lista}

            for nome in non_convocati_lista:
                col_nome, col_motivo = st.columns([1, 2])
                with col_nome:
                    st.text(nome)
                with col_motivo:
                    st.session_state.motivi_non_convocati[nome] = st.selectbox(
                        "Motivo",
                        options=[""] + motivi_disponibili,
                        index=motivi_disponibili.index(st.session_state.motivi_non_convocati[nome]) if st.session_state.motivi_non_convocati[nome] in motivi_disponibili else 0,
                        key=f"motivo_nonconv_{nome}"
                    )

            st.markdown("---")
            if st.button("💾 Salva partita"):
                squadra_nome = squadra_avversaria.replace(" ", "_").upper()
                nome_file = f"{giornata}_{squadra_nome}.json"
                dir_partita_squadra = os.path.join("partita", squadra_sel)
                os.makedirs(dir_partita_squadra, exist_ok=True)
                path_file = os.path.join(dir_partita_squadra, nome_file)

                dati_partita = {
                    "giornata": giornata,
                    "squadra": squadra_nome,
                    "home_away": home_away,
                    "risultato": risultato,
                    "recupero": recupero,
                    "formazione": st.session_state.formazione,
                    "substitutions": [
                        {
                            "sub_in": s["in"],
                            "sub_out": s["out"],
                            "time_sub": int(s["minuto"])
                        }
                        for s in st.session_state.sostituzioni
                        if s["in"] and s["out"] and s["minuto"].isdigit()
                    ],
                    "ammonizioni": st.session_state.ammoniti,
                    "espulsioni": [
                        {
                            "esp_player": e["giocatore"],
                            "time_esp": int(e["minuto"])
                        }
                        for e in st.session_state.espulsioni
                        if e["giocatore"] and e["minuto"].isdigit()
                    ],
                    "goal": st.session_state.gol,
                    "non_convocati": [
                        {
                            "giocatore": nome,
                            "motivo": st.session_state.motivi_non_convocati.get(nome, "").upper()
                        }
                        for nome in st.session_state.motivi_non_convocati
                        if nome
                    ]
                }

                with open(path_file, "w", encoding="utf-8") as f:
                    json.dump(dati_partita, f, ensure_ascii=False, indent=2)

                st.success(f"File JSON salvato in: {path_file}")

                pdf_buffer = genera_pdf_partita(dati_partita)
                st.download_button(
                    label="📄 Scarica Report PDF",
                    data=pdf_buffer,
                    file_name=nome_file.replace(".json", ".pdf"),
                    mime="application/pdf"
                )

    elif st.session_state.sezione == "Reportistica":
        from calculate_minutes import main as calculate_minutes, load_match_data

        st.markdown(f"### 📊 Statistiche Aggregate – Squadra **{squadra_sel}**")

        dir_partite = os.path.join("partita", squadra_sel)
        partite_files = [f for f in os.listdir(dir_partite) if f.endswith(".json")] if os.path.exists(dir_partite) else []

        if not partite_files:
            st.warning("⚠️ Nessuna partita trovata per questa squadra.")
        else:
            # Inizializza strutture per le statistiche
            total_player_minutes = {}
            total_player_starts = {}
            total_player_subs_in = {}
            total_player_subs_out = {}
            total_player_yellow_cards = {}
            total_player_red_cards = {}
            total_player_goals = {}
            total_player_matches = {}

            total_yellow_cards = 0
            total_red_cards = 0
            total_goals = 0
            matches_played = 0
            total_goals_conceded = 0

            # Durata partita in base alla categoria
            if squadra_sel in ["PP", "U19", "U18"] or squadra_sel.startswith("U17"):
                durata_partita = 90
            elif squadra_sel.startswith("U16"):
                durata_partita = 80
            elif squadra_sel.startswith("U15") or squadra_sel.startswith("U14"):
                durata_partita = 70
            else:
                durata_partita = 0  # fallback


            for partita_file in partite_files:
                try:
                    partita_path = os.path.join(dir_partite, partita_file)
                    with open(partita_path, 'r') as f:
                        match_data = json.load(f)
                    
                    # Estrai i gol subiti dalla stringa risultato (es: "2-1" → prende "1")
                    try:
                        risultato = match_data.get("risultato", "0-0")
                        gol_subiti = int(risultato.split("-")[1].strip())
                    except Exception:
                        gol_subiti = 0
                    total_goals_conceded += gol_subiti

                    player_minutes, player_status, _ = calculate_minutes(partita_path)
                    matches_played += 1

                    for player, status in player_status.items():
                        if player not in total_player_minutes:
                            total_player_minutes[player] = 0
                            total_player_starts[player] = 0
                            total_player_subs_in[player] = 0
                            total_player_subs_out[player] = 0
                            total_player_yellow_cards[player] = 0
                            total_player_red_cards[player] = 0
                            total_player_goals[player] = 0
                            total_player_matches[player] = 0

                        if player in player_minutes:
                            total_player_minutes[player] += player_minutes[player]
                            if player_minutes[player] > 0:
                                total_player_matches[player] += 1

                        status_parts = status.split(" | ")

                        if "Titolare" in status_parts:
                            total_player_starts[player] += 1
                        if "Sostituito" in status_parts:
                            total_player_subs_out[player] += 1
                        if "Subentrato" in status_parts:
                            total_player_subs_in[player] += 1
                        if "Ammonito" in status_parts:
                            total_player_yellow_cards[player] += 1
                            total_yellow_cards += 1
                        if "Espulso" in status_parts:
                            total_player_red_cards[player] += 1
                            total_red_cards += 1
                        if "Gol" in status_parts:
                            total_player_goals[player] += 1
                            total_goals += 1

                except Exception as e:
                    st.error(f"❌ Errore processando il file {partita_file}: {str(e)}")

            player_avg_minutes = {
                player: (total_player_minutes[player] / total_player_matches[player]) if total_player_matches[player] > 0 else 0
                for player in total_player_minutes
            }

            player_stats = [{
                'Giocatore': player,
                'Partite': total_player_matches[player],
                'Minuti': total_player_minutes[player],
                'Minuti Disponibili': total_player_matches[player] * durata_partita,
                'Media Minuti': round(player_avg_minutes[player], 1),
                'Titolari': total_player_starts[player],
                'Subentri': total_player_subs_in[player],
                'Sostituzioni': total_player_subs_out[player],
                'Gol': total_player_goals[player],
                'Ammonizioni': total_player_yellow_cards[player],
                'Espulsioni': total_player_red_cards[player]
            } for player in total_player_minutes]


            player_stats = sorted(player_stats, key=lambda x: x['Minuti'], reverse=True)

            st.divider()
            st.markdown("### 📈 Statistiche Generali")

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("📅 Partite giocate", matches_played)
                st.metric("⚽ Gol totali", total_goals)
            with col2:
                st.metric("🟨 Ammonizioni", total_yellow_cards)
                st.metric("📊 Media gol/partita", round(total_goals/matches_played, 1) if matches_played > 0 else 0)
            with col3:
                st.metric("🟥 Espulsioni", total_red_cards)
                st.metric("⚽ Gol subiti", total_goals_conceded)

            st.divider()
            st.markdown("### 👥 Statistiche per Giocatore")

            df_stats = pd.DataFrame(player_stats)
            styled_df = df_stats.style.format({'Media Minuti': '{:.1f}'}).background_gradient(subset='Minuti', cmap='Blues')

            st.dataframe(
                styled_df,
                use_container_width=True,
                hide_index=True
            )

            st.divider()

            st.markdown("### 🕸️ Confronto Giocatori (Grafico Radar)")

            giocatori_selezionati = st.multiselect(
                "Seleziona i giocatori da confrontare:",
                options=[p["Giocatore"] for p in player_stats],
                default=[],
                help="Puoi selezionare da 2 a 5 giocatori"
            )

            visualizza_percentuale = st.toggle("Visualizza in percentuale")

            if 2 <= len(giocatori_selezionati) <= 5:
                metriche = ["Minuti", "Titolari", "Subentri", "Sostituzioni"]
                fig = go.Figure()

                for nome in giocatori_selezionati:
                    g = next((p for p in player_stats if p["Giocatore"] == nome), None)
                    if g:
                        if visualizza_percentuale:
                            # Calcolo percentuali
                            minuti_disp = g["Partite"] * durata_partita if durata_partita > 0 else 1
                            minuti_pct = (g["Minuti"] / minuti_disp) * 100 if minuti_disp else 0
                            titolari_pct = (g["Titolari"] / g["Partite"]) * 100 if g["Partite"] else 0
                            subentri_pct = (g["Subentri"] / g["Partite"]) * 100 if g["Partite"] else 0
                            sostituzioni_pct = (g["Sostituzioni"] / g["Partite"]) * 100 if g["Partite"] else 0
                            valori = [minuti_pct, titolari_pct, subentri_pct, sostituzioni_pct]
                        else:
                            # Valori assoluti
                            valori = [g["Minuti"], g["Titolari"], g["Subentri"], g["Sostituzioni"]]

                        fig.add_trace(go.Scatterpolar(
                            r=valori,
                            theta=metriche,
                            fill='toself',
                            name=nome
                        ))

                # Configura grafico radar
                fig.update_layout(
                    polar=dict(
                        radialaxis=dict(
                            visible=True,
                            range=[0, 100] if visualizza_percentuale else None
                        )
                    ),
                    title="Confronto Giocatori (Percentuali)" if visualizza_percentuale else "Confronto Giocatori (Valori Assoluti)",
                    showlegend=True,
                    height=500
                )

                st.plotly_chart(fig, use_container_width=True)

            elif len(giocatori_selezionati) == 1:
                st.info("⚠️ Seleziona almeno 2 giocatori per visualizzare il grafico radar.")
            elif len(giocatori_selezionati) > 5:
                st.warning("⚠️ Puoi selezionare al massimo 5 giocatori per questo confronto.")


            st.divider()

            st.markdown("### 📊 Grafici")

            # Scelta tipo di grafico
            grafico_scelto = st.selectbox(
                "Scegli il tipo di grafico da visualizzare:",
                [
                    "⏱️ Minuti Giocati vs Minuti Disponibili",
                    "🧤 Titolare vs Totale Partite",
                    "🔁 Subentrato vs Totale Partite"
                ]
            )

            # Crea DataFrame base
            df_plot = pd.DataFrame(player_stats)
            df_plot["Giocatore"] = df_plot["Giocatore"].fillna("Sconosciuto")

            # Visualizzazione grafico scelto
            if grafico_scelto == "⏱️ Minuti Giocati vs Minuti Disponibili":
                fig = px.bar(
                    df_plot,
                    x="Giocatore",
                    y=["Minuti", "Minuti Disponibili"],
                    barmode="group",
                    labels={"value": "Minuti", "variable": "Tipo"},
                    title="Minuti Giocati vs Minuti Disponibili",
                    color_discrete_map={"Minuti": "steelblue", "Minuti Disponibili": "lightgray"},
                    height=500
                )

            elif grafico_scelto == "🧤 Titolare vs Totale Partite":
                fig = px.bar(
                    df_plot,
                    x="Giocatore",
                    y=["Titolari", "Partite"],
                    barmode="group",
                    labels={"value": "Numero", "variable": "Tipo"},
                    title="Partite da Titolare vs Totali",
                    color_discrete_map={"Titolari": "green", "Partite": "gray"},
                    height=500
                )

            elif grafico_scelto == "🔁 Subentrato vs Totale Partite":
                fig = px.bar(
                    df_plot,
                    x="Giocatore",
                    y=["Subentri", "Partite"],
                    barmode="group",
                    labels={"value": "Numero", "variable": "Tipo"},
                    title="Subentri vs Partite Giocate",
                    color_discrete_map={"Subentri": "orange", "Partite": "gray"},
                    height=500
                )

            fig.update_layout(xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)


